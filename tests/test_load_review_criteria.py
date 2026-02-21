from pathlib import Path

import numpy as np
import pytest
from openpyxl import Workbook

from contract_review.criteria_parser import parse_criteria_excel
from contract_review.skills.local.load_review_criteria import (
    LoadReviewCriteriaInput,
    _normalize_clause_ref,
    load_review_criteria,
)


@pytest.fixture
def sample_criteria_xlsx(tmp_path: Path) -> Path:
    wb = Workbook()
    ws = wb.active
    ws.append(["条款编号", "条款名称", "审核要点", "风险等级", "标准条件", "建议措施"])
    ws.append(["4.1", "承包商义务", "义务范围不应超出原文", "高", "GC 原文", "建议限缩"])
    ws.append(["20.1", "承包商索赔", "通知期限不应短于28天", "高", "28天", "建议恢复"])
    ws.append(["14.7", "期中付款", "付款周期应合理", "中", "56天", "建议缩短"])
    path = tmp_path / "criteria.xlsx"
    wb.save(path)
    return path


def test_parse_criteria_excel_normal(sample_criteria_xlsx: Path):
    rows = parse_criteria_excel(sample_criteria_xlsx)
    assert len(rows) == 3
    assert rows[0].criterion_id == "RC-1"
    assert rows[0].clause_ref == "4.1"


def test_parse_criteria_excel_column_recognition(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.append(["Clause", "Title", "Check Point", "Risk Level", "Benchmark", "Recommendation"])
    ws.append(["4.1", "Obligations", "Scope should remain balanced", "high", "GC baseline", "narrow obligations"])
    path = tmp_path / "criteria_en.xlsx"
    wb.save(path)
    rows = parse_criteria_excel(path)
    assert len(rows) == 1
    assert rows[0].review_point == "Scope should remain balanced"


def test_parse_criteria_excel_empty_rows(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.append(["条款编号", "审核要点"])
    ws.append([None, None])
    ws.append(["4.1", "义务范围不应超出原文"])
    path = tmp_path / "criteria_empty.xlsx"
    wb.save(path)
    rows = parse_criteria_excel(path)
    assert len(rows) == 1


def test_parse_criteria_excel_missing_columns(tmp_path: Path):
    wb = Workbook()
    ws = wb.active
    ws.append(["条款编号", "审核要点"])
    ws.append(["4.1", "义务范围不应超出原文"])
    path = tmp_path / "criteria_missing.xlsx"
    wb.save(path)
    rows = parse_criteria_excel(path)
    assert len(rows) == 1
    assert rows[0].risk_level == ""
    assert rows[0].baseline_text == ""


def test_parse_criteria_excel_file_not_found(tmp_path: Path):
    rows = parse_criteria_excel(tmp_path / "not_found.xlsx")
    assert rows == []


@pytest.mark.asyncio
async def test_load_criteria_exact_match(sample_criteria_xlsx: Path):
    criteria = [row.model_dump() for row in parse_criteria_excel(sample_criteria_xlsx)]
    result = await load_review_criteria(
        LoadReviewCriteriaInput(
            clause_id="4.1",
            document_structure={"clauses": []},
            criteria_data=criteria,
        )
    )
    assert result.has_criteria is True
    assert result.total_matched >= 1
    assert result.matched_criteria[0].match_type == "exact"


@pytest.mark.asyncio
async def test_load_criteria_no_criteria():
    result = await load_review_criteria(
        LoadReviewCriteriaInput(
            clause_id="4.1",
            document_structure={"clauses": []},
        )
    )
    assert result.has_criteria is False
    assert result.total_matched == 0


@pytest.mark.asyncio
async def test_load_criteria_no_match(sample_criteria_xlsx: Path):
    criteria = [row.model_dump() for row in parse_criteria_excel(sample_criteria_xlsx)]
    result = await load_review_criteria(
        LoadReviewCriteriaInput(
            clause_id="99.9",
            document_structure={"clauses": []},
            criteria_data=criteria,
        )
    )
    assert result.has_criteria is True
    assert result.total_matched == 0


@pytest.mark.asyncio
async def test_load_criteria_semantic_fallback(monkeypatch, sample_criteria_xlsx: Path):
    criteria = [row.model_dump() for row in parse_criteria_excel(sample_criteria_xlsx)]
    monkeypatch.setattr(
        "contract_review.skills.local.load_review_criteria._embed_texts",
        lambda _texts: np.array(
            [
                [1.0, 0.0],
                [0.2, 0.9],
                [0.95, 0.0],
                [0.3, 0.8],
            ]
        ),
    )
    result = await load_review_criteria(
        LoadReviewCriteriaInput(
            clause_id="8.8",
            document_structure={"clauses": [{"clause_id": "8.8", "text": "通知期限不应短于28天", "children": []}]},
            criteria_data=criteria,
        )
    )
    assert result.has_criteria is True
    assert result.total_matched >= 1
    assert result.matched_criteria[0].match_type == "semantic"


def test_load_criteria_normalize_clause_ref():
    assert _normalize_clause_ref("Sub-Clause 4.1") == "4.1"
    assert _normalize_clause_ref("条款 4.1 ") == "4.1"
    assert _normalize_clause_ref("Clause 20.1.") == "20.1"
