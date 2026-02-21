"""Excel parser for review criteria tables."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook

from .models import ReviewCriterion

logger = logging.getLogger(__name__)


COLUMN_ROLE_KEYWORDS: Dict[str, List[str]] = {
    "clause_ref": ["条款编号", "条款号", "clause", "sub-clause", "编号", "ref"],
    "clause_name": ["条款名称", "名称", "clause name", "title", "标题"],
    "review_point": ["审核要点", "审查要点", "要点", "review point", "check point", "审核内容", "检查内容", "issue"],
    "risk_level": ["风险等级", "风险", "等级", "risk", "level", "severity", "优先级", "priority"],
    "baseline_text": ["标准条件", "基准", "baseline", "standard", "benchmark", "参考标准", "原文"],
    "suggested_action": ["建议措施", "建议", "措施", "suggestion", "action", "recommendation", "应对", "对策"],
}


def _normalize_header(text: str) -> str:
    return "".join((text or "").strip().lower().split())


def _resolve_column_roles(headers: list[str]) -> Dict[str, int]:
    role_map: Dict[str, int] = {}
    normalized = [_normalize_header(h) for h in headers]
    for idx, col in enumerate(normalized):
        if not col:
            continue
        for role, keywords in COLUMN_ROLE_KEYWORDS.items():
            if role in role_map:
                continue
            for keyword in keywords:
                if _normalize_header(keyword) in col:
                    role_map[role] = idx
                    break
    return role_map


def _fallback_review_point_column(rows: list[list[str]], role_map: Dict[str, int], headers: list[str]) -> Dict[str, int]:
    if "review_point" in role_map:
        return role_map
    if not rows:
        return role_map
    best_idx = -1
    best_score = -1
    for idx in range(len(headers)):
        values = [str(row[idx]).strip() for row in rows if idx < len(row) and str(row[idx]).strip()]
        if not values:
            continue
        score = sum(len(v) for v in values)
        if score > best_score:
            best_idx = idx
            best_score = score
    if best_idx >= 0:
        role_map["review_point"] = best_idx
        logger.warning("未识别到审核要点列，已回退到第 %d 列", best_idx + 1)
    return role_map


def parse_criteria_excel(file_path: str | Path, sheet_name: str | None = None) -> list[ReviewCriterion]:
    path = Path(file_path)
    if not path.exists():
        logger.warning("审核标准文件不存在: %s", path)
        return []
    try:
        wb = load_workbook(path, data_only=True)
    except Exception as exc:
        logger.warning("读取审核标准 Excel 失败: %s", exc)
        return []

    try:
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
    except Exception:
        return []

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    if not any(headers):
        return []

    data_rows = []
    for row in rows[1:]:
        values = ["" if v is None else str(v).strip() for v in row]
        if any(values):
            data_rows.append(values)
    if not data_rows:
        return []

    role_map = _resolve_column_roles(headers)
    role_map = _fallback_review_point_column(data_rows, role_map, headers)

    criteria: list[ReviewCriterion] = []
    for idx, row in enumerate(data_rows, start=1):
        raw = {headers[col_idx]: (row[col_idx] if col_idx < len(row) else "") for col_idx in range(len(headers))}

        def pick(role: str) -> str:
            col_idx = role_map.get(role)
            if col_idx is None or col_idx >= len(row):
                return ""
            return str(row[col_idx]).strip()

        review_point = pick("review_point")
        if not review_point:
            continue
        criteria.append(
            ReviewCriterion(
                criterion_id=f"RC-{idx}",
                clause_ref=pick("clause_ref"),
                clause_name=pick("clause_name"),
                review_point=review_point,
                risk_level=pick("risk_level"),
                baseline_text=pick("baseline_text"),
                suggested_action=pick("suggested_action"),
                raw_row=raw,
            )
        )
    return criteria
